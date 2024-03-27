import pytest
import time
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

@pytest.fixture(scope="session")
def browser():
    driver = webdriver.Chrome()
    yield driver
    driver.quit()

# @pytest.fixture(scope="module")
def test_data():
    workbook = load_workbook(filename="DDT-Selenium\\exelsheet\\TestCasesFramework2.xlsx")
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=6, values_only=True):
        action = row[0]  # Column 3
        data_value = row[1]  # Column 4
        locator = row[2]  # Column 5
        locator_type = row[3]  # Column 6
        data.append((action, data_value, locator, locator_type))
    workbook.close()
    return data

@pytest.mark.parametrize("action, data, locator, locator_type", test_data())
def test_excel_actions(browser, action, data, locator, locator_type):
    if action == "Navigate":
        browser.get(data)
        time.sleep(2)
    elif action == "Click":
        if locator_type == "class":
            element = browser.find_element(By.CLASS_NAME, locator)
            element.click()
            time.sleep(2)
        elif locator_type == "xpath":
            element = browser.find_element(By.XPATH, locator)
            element.click()
            time.sleep(2)
        elif locator_type == "id":
            element = browser.find_element(By.ID, locator)
            element.click()
            time.sleep(2)
    elif action == "Type":
        if locator_type == "class":
            element = browser.find_element(By.CLASS_NAME, locator)
            element.send_keys(data)
            time.sleep(2)
        elif locator_type == "xpath":
            element = browser.find_element(By.XPATH, locator)
            element.send_keys(data)
            time.sleep(2)
        elif locator_type == "id":
            element = browser.find_element(By.ID, locator)
            element.send_keys(data)
            time.sleep(2)
    elif action == "Enter":  # Check if action is "Enter"
        if locator_type == "class":
            element = browser.find_element(By.CLASS_NAME, locator)
            element.send_keys(Keys.ENTER)
            time.sleep(2)
        elif locator_type == "xpath":
            element = browser.find_element(By.XPATH, locator)
            element.send_keys(Keys.ENTER)
            time.sleep(2)
        elif locator_type == "id":
            element = browser.find_element(By.ID, locator)
            element.send_keys(Keys.ENTER)
            time.sleep(2)
